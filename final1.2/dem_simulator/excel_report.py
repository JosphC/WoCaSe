"""Excel report generation (mirrors results wcs.docx)."""

from __future__ import annotations

import logging
import os
import time
from typing import Any, Dict, List, Optional, Tuple

from dem_simulator.exceptions import ExcelReportError
from dem_simulator.constants import (
    MC_DEFAULT_CYCLES,
    MC_DEFAULT_SEED,
    SENSITIVITY_POST_VALS,
    SENSITIVITY_EA_VALS,
)
from dem_simulator.config import (
    ProjectConfig,
    ProjectDefinition,
)
from dem_simulator.costs import MicroCosts, COSTS_PROJ2, COSTS_PROJ3
from dem_simulator.scenarios import (
    WCS_SCENARIO_KEYS,
    build_scenarios,
)
from dem_simulator.config import DEFAULT_CALIBRATIONS
from dem_simulator.engine import DemMainFunctionSimulator
from dem_simulator.simulation import (
    simulate_wcs_grid,
    compute_rmse,
    compute_per_scenario_rmse,
)
from dem_simulator.analysis import simulate_peak_runtime

logger = logging.getLogger("dem_simulator")

# Lazy import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference as ChartRef
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ==============================================================================
# Helpers
# ==============================================================================

def _apply_table_border(ws: Any, min_row: int, max_row: int,
                        min_col: int, max_col: int) -> None:
    """Apply thin borders to a rectangular cell range."""
    thin = Side(style="thin")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=thin, right=thin, top=thin, bottom=thin,
            )


def _set_column_widths(ws: Any, widths: Dict[str, float]) -> None:
    """Set column widths from a mapping ``{ column_letter: width }``."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ==============================================================================
# Main report generator
# ==============================================================================

def generate_excel_report(
    cfg: ProjectConfig,
    costs: MicroCosts,
    output_path: str = "results_wcs_simulated.xlsx",
    project: Optional[ProjectDefinition] = None,
    *,
    include_monte_carlo: bool = True,
    mc_cycles: int = MC_DEFAULT_CYCLES,
    mc_seed: int = MC_DEFAULT_SEED,
    version: str = "",
    yellow_threshold: int = 500,
) -> str:
    """Generate an Excel workbook whose structure mirrors *results wcs.docx*.

    Sheets
    ------
    1. WCS Results      – main grid (3 scenarios × 6 calibrations)
    2. Comparison       – simulated vs bench reference (if available)
    3. Breakdown        – detailed per-phase cost breakdown
    4. Sensitivity      – NrClcFmyEveAsyn × NrClcFmyPost sweep
    5. Cross-Variant    – comparison across PTU variants
    6. Fitted Costs     – PROJ2 vs PROJ3 cost comparison
    7. Monte Carlo      – stochastic runtime distribution (optional)

    Parameters
    ----------
    cfg : ProjectConfig
        The project configuration.
    costs : MicroCosts
        Fitted micro-costs.
    output_path : str
        File path for the generated ``.xlsx`` file.
    project : ProjectDefinition, optional
        Full project definition with reference data.
    include_monte_carlo : bool
        Whether to add the Monte Carlo sheet.
    mc_cycles : int
        Number of Monte Carlo cycles.
    mc_seed : int
        RNG seed for reproducibility.
    version : str
        Simulator version string shown in the report.

    Returns
    -------
    str
        Absolute path of the saved workbook.

    Raises
    ------
    ExcelReportError
        If ``openpyxl`` is not installed or the file cannot be saved.
    """
    if not HAS_OPENPYXL:
        raise ExcelReportError(
            "openpyxl is required for Excel report generation.  "
            "Install it with:  pip install openpyxl"
        )

    # Import version from package if not provided
    if not version:
        try:
            from dem_simulator import __version__
            version = __version__
        except ImportError:
            version = "unknown"

    proj_name = project.name if project else cfg.name
    ref_wcs = project.reference_wcs if project else {}
    has_reference = bool(ref_wcs)
    calibrations = project.calibrations if project else DEFAULT_CALIBRATIONS

    logger.info("Generating Excel report for %s ...", proj_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "WCS Results"

    # ── Styles ────────────────────────────────────────────────────────────────
    bold = Font(bold=True)
    bold_blue = Font(bold=True, color="0000FF")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                              fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                             fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="center")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Table 1: Configuration Info ───────────────────────────────────────────
    config_rows: List[Tuple[str, str, str]] = [
        ("Simulator Version:", version, ""),
        ("Project:", proj_name, project.description if project else ""),
        ("Final calibration value for serial production:", "", ""),
        ("Value in project:", str(cfg.NrClcFmyEveAsyn), ""),
        ("Description:", "", ""),
        (
            "NrFmy in project:",
            str(cfg.NrFmy),
            "NrFmy configured in Icsp_dem_genr_cnf_spec.arxml "
            "-> DemGenr -> Fmy -> NrFmy",
        ),
        (
            "Icsp_Dem_Fmy_NrClcFmyEveAsyn:",
            str(cfg.NrClcFmyEveAsyn),
            "Maximum number of event reports to be processed "
            "in one recurrence",
        ),
        (
            "Icsp_Dem_Fmy_NrClcFmyPost:",
            str(cfg.NrClcFmyPost),
            "Maximum number of failure memory entries to be "
            "processed in one recurrence (freeze-frames and "
            "callbacks)",
        ),
        ("NrEve:", str(cfg.NrEve), "Total number of DEM events"),
        ("NrFrfDataTot:", str(cfg.NrFrfDataTot),
         "Total freeze-frame data elements"),
        ("NrLamp:", str(cfg.NrLamp), "Number of lamp entries"),
        ("NrFrfPre:", str(cfg.NrFrfPre), "Prestored freeze frames"),
        ("NrBlockFrf:", str(cfg.NrBlockFrf), "Number of FRF blocks"),
        ("MBT (maximum blocking time) - confirmed by project:", "-", ""),
        ("Core in project where Dem_MainFunction is located:", "-", ""),
        ("Task in project where Dem_MainFunction is called:", "-", ""),
    ]

    for ri, (c1, c2, c3) in enumerate(config_rows, start=1):
        ws.cell(row=ri, column=1, value=c1).font = bold
        ws.cell(row=ri, column=2, value=c2)
        ws.cell(row=ri, column=3, value=c3).alignment = wrap

    _apply_table_border(ws, 1, len(config_rows), 1, 3)

    # ── Blank row + project heading ───────────────────────────────────────────
    gap_row = len(config_rows) + 2
    ws.cell(row=gap_row, column=1, value=proj_name).font = Font(
        bold=True, size=14,
    )

    # ── Table 2: WCS Results Grid ─────────────────────────────────────────────
    tbl_start = gap_row + 1

    headers = [
        "Worst case scenarios:",
        "NrClcFmyEveAsyn / NrClcFmyPost"
    ]
    for i, (asyn, post) in enumerate(calibrations):
        headers.append(f"Calibration {i + 1}: {asyn}/{post}")

    for ci, hdr_text in enumerate(headers, start=1):
        cell = ws.cell(row=tbl_start, column=ci, value=hdr_text)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    # ── Simulate WCS grid ─────────────────────────────────────────────────────
    grid = simulate_wcs_grid(cfg, costs, calibrations)
    scenarios = build_scenarios(
        nr_inject_first=cfg.nr_inject_first,
        nr_inject_next=cfg.nr_inject_next,
        nr_fmy=cfg.NrFmy,
    )

    for si, scenario in enumerate(scenarios):
        skey = WCS_SCENARIO_KEYS[si]
        row_num = tbl_start + 1 + si
        label = f"{scenario.name}:\u00a0{scenario.description}"
        ws.cell(row=row_num, column=1, value=label).alignment = wrap
        ws.cell(row=row_num, column=2, value="")

        sim_values = grid[skey]
        for ci, sim_val in enumerate(sim_values):
            cell = ws.cell(row=row_num, column=3 + ci)
            cell.alignment = center
            if sim_val >= yellow_threshold:
                cell.value = f"{sim_val}\u00a0"
                cell.fill = yellow_fill
            else:
                cell.value = str(sim_val)
                cell.fill = green_fill

    tbl_end_row = tbl_start + len(scenarios)
    _apply_table_border(ws, tbl_start, tbl_end_row, 1, 2 + len(calibrations))

    # ── Comparison sheet (only if reference data exists) ──────────────────────
    if has_reference:
        ws2 = wb.create_sheet("Comparison")
        ws2.cell(
            row=1, column=1,
            value=f"Simulated vs Reference ({proj_name})",
        ).font = Font(bold=True, size=14)

        rmse = compute_rmse(cfg, costs, ref_wcs, calibrations)
        ws2.cell(row=2, column=1,
                 value=f"Overall RMSE: {rmse:.1f} \u00b5s").font = bold

        comp_headers = [
            "Scenario", "Calibration", "Reference (\u00b5s)",
            "Simulated (\u00b5s)", "Diff (\u00b5s)", "Diff (%)",
        ]
        for ci, h in enumerate(comp_headers, start=1):
            cell = ws2.cell(row=4, column=ci, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center

        row_num = 5
        for si, scenario in enumerate(scenarios):
            skey = WCS_SCENARIO_KEYS[si]
            sim_values = grid[skey]
            ref_values = ref_wcs.get(skey, [0] * len(calibrations))
            for _ci2, ((asyn, post), sim_v, ref_v) in enumerate(
                zip(calibrations, sim_values, ref_values)
            ):
                ws2.cell(row=row_num, column=1, value=scenario.name)
                ws2.cell(row=row_num, column=2, value=f"{asyn}/{post}")
                ws2.cell(row=row_num, column=3, value=ref_v).alignment = center
                ws2.cell(row=row_num, column=4, value=sim_v).alignment = center
                diff_abs = sim_v - ref_v
                ws2.cell(row=row_num, column=5,
                         value=round(diff_abs, 1)).alignment = center
                diff_pct = (
                    ((sim_v - ref_v) / ref_v) * 100 if ref_v else 0.0
                )
                diff_cell = ws2.cell(
                    row=row_num, column=6, value=round(diff_pct, 1),
                )
                diff_cell.alignment = center
                diff_cell.number_format = '0.0"%"'
                if abs(diff_pct) < 10:
                    diff_cell.fill = green_fill
                elif abs(diff_pct) >= 15:
                    diff_cell.fill = yellow_fill
                row_num += 1

        _apply_table_border(ws2, 4, row_num - 1, 1, 6)
        _set_column_widths(ws2, {"A": 35, "B": 14})
        for c in range(3, 7):
            ws2.column_dimensions[get_column_letter(c)].width = 16

        # Per-scenario RMSE summary
        per_rmse = compute_per_scenario_rmse(cfg, costs, ref_wcs, calibrations)
        rmse_row = row_num + 1
        ws2.cell(row=rmse_row, column=1,
                 value="Per-Scenario RMSE (\u00b5s)").font = bold
        for si, skey in enumerate(WCS_SCENARIO_KEYS):
            ws2.cell(row=rmse_row + 1 + si, column=1, value=skey)
            ws2.cell(row=rmse_row + 1 + si, column=2,
                     value=round(per_rmse.get(skey, 0.0), 1)).alignment = center

        # Bar chart: simulated vs reference for calibration 1
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = f"Sim vs Ref – Cal1 ({proj_name})"
            chart.y_axis.title = "Runtime (\u00b5s)"
            chart.x_axis.title = "Scenario"

            # Build a small data table for the chart
            ch_start_row = rmse_row + len(WCS_SCENARIO_KEYS) + 3
            ws2.cell(row=ch_start_row, column=1, value="Scenario")
            ws2.cell(row=ch_start_row, column=2, value="Reference")
            ws2.cell(row=ch_start_row, column=3, value="Simulated")
            for si, skey in enumerate(WCS_SCENARIO_KEYS):
                ws2.cell(row=ch_start_row + 1 + si, column=1, value=skey)
                ref_vals = ref_wcs.get(skey, [0])
                ws2.cell(row=ch_start_row + 1 + si, column=2,
                         value=ref_vals[0] if ref_vals else 0)
                ws2.cell(row=ch_start_row + 1 + si, column=3,
                         value=grid[skey][0])

            data_ref = ChartRef(ws2, min_col=2, max_col=3,
                                min_row=ch_start_row,
                                max_row=ch_start_row + len(WCS_SCENARIO_KEYS))
            cats = ChartRef(ws2, min_col=1, min_row=ch_start_row + 1,
                            max_row=ch_start_row + len(WCS_SCENARIO_KEYS))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 18
            chart.height = 10
            ws2.add_chart(chart, f"E{rmse_row}")
        except Exception:  # noqa: BLE001
            logger.warning("Could not create bar chart in Comparison sheet.")

    # ── Detailed Breakdown sheet ──────────────────────────────────────────────
    ws_bd = wb.create_sheet("Breakdown")
    ws_bd.cell(
        row=1, column=1,
        value=f"Detailed Cost Breakdown ({proj_name})",
    ).font = Font(bold=True, size=14)

    sim_obj = DemMainFunctionSimulator(cfg, costs)
    bd_headers = ["Phase"] + [
        f"Cal{i + 1} ({a}/{p})"
        for i, (a, p) in enumerate(calibrations)
    ]
    for ci, h in enumerate(bd_headers, start=1):
        cell = ws_bd.cell(row=3, column=ci, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    row_num = 4
    for _si, scenario in enumerate(scenarios):
        ws_bd.cell(row=row_num, column=1, value=scenario.name).font = bold_blue
        row_num += 1

        all_bds: List[Dict[str, float]] = []
        for asyn, post in calibrations:
            bd = sim_obj.simulate(scenario, asyn, post)
            all_bds.append(bd)
        phases = list(all_bds[0].keys())

        for phase in phases:
            ws_bd.cell(row=row_num, column=1, value=phase)
            for ci, bd in enumerate(all_bds):
                cell = ws_bd.cell(row=row_num, column=2 + ci,
                                  value=round(bd[phase], 2))
                cell.alignment = center
                cell.number_format = "0.00"
            row_num += 1

        # Total row
        ws_bd.cell(row=row_num, column=1, value="TOTAL").font = bold
        for ci, bd in enumerate(all_bds):
            total = sum(bd.values())
            cell = ws_bd.cell(row=row_num, column=2 + ci,
                              value=round(total, 1))
            cell.alignment = center
            cell.font = bold
            cell.number_format = "0.0"
        row_num += 2

    ws_bd.column_dimensions["A"].width = 22
    for c in range(2, 2 + len(calibrations)):
        ws_bd.column_dimensions[get_column_letter(c)].width = 16

    # ── Sensitivity Analysis sheet ────────────────────────────────────────────
    ws3 = wb.create_sheet("Sensitivity")
    ws3.cell(
        row=1, column=1,
        value=f"Parameter Sensitivity \u2014 Sweep ({proj_name})",
    ).font = Font(bold=True, size=14)

    post_vals = list(SENSITIVITY_POST_VALS)
    ea_vals = list(SENSITIVITY_EA_VALS)

    row_num = 3
    for _si, scenario in enumerate(scenarios):
        ws3.cell(row=row_num, column=1, value=scenario.name).font = bold_blue
        row_num += 1

        ws3.cell(row=row_num, column=1, value="EveAsyn \\ Post").font = bold
        for ci, p in enumerate(post_vals):
            cell = ws3.cell(row=row_num, column=2 + ci, value=f"Post={p}")
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
        row_num += 1

        for ea in ea_vals:
            ws3.cell(row=row_num, column=1,
                     value=f"EveAsyn={ea}").font = bold
            for ci, fp in enumerate(post_vals):
                t = sim_obj.simulate_total(scenario, ea, fp)
                cell = ws3.cell(row=row_num, column=2 + ci, value=round(t))
                cell.alignment = center
                if t >= yellow_threshold:
                    cell.fill = yellow_fill
            row_num += 1
        row_num += 1

    ws3.column_dimensions["A"].width = 20
    for c in range(2, 2 + len(post_vals)):
        ws3.column_dimensions[get_column_letter(c)].width = 12

    # ── Cross-Variant sheet ───────────────────────────────────────────────────
    variants = project.variant_configs if project else {}
    ws4 = wb.create_sheet("Cross-Variant")
    ws4.cell(
        row=1, column=1,
        value=f"Cross-Variant Comparison ({proj_name})",
    ).font = Font(bold=True, size=14)

    cv_headers = [
        "Variant", "NrFmy", "NrEve", "NrClient",
        "S1 (\u00b5s)", "S2 (\u00b5s)", "S3 (\u00b5s)",
    ]
    for ci, h in enumerate(cv_headers, start=1):
        cell = ws4.cell(row=3, column=ci, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    row_num = 4
    for vname, vcfg in sorted(variants.items()):
        vsim = DemMainFunctionSimulator(vcfg, costs)
        ws4.cell(row=row_num, column=1, value=vname)
        ws4.cell(row=row_num, column=2,
                 value=vcfg.NrFmy).alignment = center
        ws4.cell(row=row_num, column=3,
                 value=vcfg.NrEve).alignment = center
        ws4.cell(row=row_num, column=4,
                 value=vcfg.NrClient).alignment = center
        v_scenarios = build_scenarios(
            nr_inject_first=vcfg.nr_inject_first,
            nr_inject_next=vcfg.nr_inject_next,
            nr_fmy=vcfg.NrFmy,
        )
        for si, scenario in enumerate(v_scenarios):
            t = vsim.simulate_total(
                scenario, vcfg.NrClcFmyEveAsyn, vcfg.NrClcFmyPost,
            )
            cell = ws4.cell(row=row_num, column=5 + si, value=round(t))
            cell.alignment = center
            if t >= yellow_threshold:
                cell.fill = yellow_fill
        row_num += 1

    _apply_table_border(ws4, 3, row_num - 1, 1, len(cv_headers))
    for c in range(1, len(cv_headers) + 1):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    # ── Fitted Costs sheet ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Fitted Costs")
    ws5.cell(
        row=1, column=1, value="Fitted Micro-Costs (\u00b5s)",
    ).font = Font(bold=True, size=14)

    cost_headers = ["Parameter", "PROJ2", "PROJ3", "Ratio (13/12)",
                    "Interpretation"]
    for ci, h in enumerate(cost_headers, start=1):
        cell = ws5.cell(row=3, column=ci, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    row_num = 4
    red_font = Font(bold=True, color="FF0000")
    for fname in sorted(vars(COSTS_PROJ2)):
        if not fname.startswith("t_"):
            continue
        v12 = getattr(COSTS_PROJ2, fname)
        v13 = getattr(COSTS_PROJ3, fname)
        ratio = v13 / v12 if v12 > 0.001 else 0.0
        ws5.cell(row=row_num, column=1, value=fname)
        ws5.cell(row=row_num, column=2,
                 value=round(v12, 3)).alignment = center
        ws5.cell(row=row_num, column=3,
                 value=round(v13, 3)).alignment = center
        ratio_cell = ws5.cell(row=row_num, column=4, value=round(ratio, 2))
        ratio_cell.alignment = center
        ratio_cell.number_format = "0.00"
        # Interpretation column
        if ratio > 1.5:
            interp = "PROJ3 significantly higher"
            ratio_cell.font = red_font
        elif ratio < 0.67:
            interp = "PROJ3 significantly lower"
            ratio_cell.font = red_font
        elif ratio > 1.1:
            interp = "PROJ3 moderately higher"
        elif ratio < 0.9:
            interp = "PROJ3 moderately lower"
        else:
            interp = "Similar"
        ws5.cell(row=row_num, column=5, value=interp)
        row_num += 1

    _apply_table_border(ws5, 3, row_num - 1, 1, 5)
    _set_column_widths(ws5, {"A": 30, "B": 14, "C": 14, "D": 14, "E": 28})

    # ── Monte Carlo sheet (optional) ──────────────────────────────────────────
    if include_monte_carlo:
        try:
            ws_mc = wb.create_sheet("Monte Carlo")
            ws_mc.cell(
                row=1, column=1,
                value=f"Monte Carlo Simulation ({proj_name})",
            ).font = Font(bold=True, size=14)

            ws_mc.cell(
                row=2, column=1,
                value=f"Cycles: {mc_cycles:,}  |  Seed: {mc_seed}",
            ).font = Font(italic=True)

            mc_headers = [
                "Statistic", "Value (\u00b5s)",
            ]
            for ci, h in enumerate(mc_headers, start=1):
                cell = ws_mc.cell(row=4, column=ci, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center

            mc_res = simulate_peak_runtime(
                cfg, costs, num_cycles=mc_cycles, seed=mc_seed,
            )

            mc_stats: List[Tuple[str, float]] = [
                ("Mean", mc_res.mean_us),
                ("Median", mc_res.median_us),
                ("Std Dev", mc_res.stdev_us),
                ("Min", mc_res.min_us),
                ("P95", mc_res.p95_us),
                ("P99", mc_res.p99_us),
                ("P99.9", mc_res.p999_us),
                ("Peak (Max)", mc_res.peak_us),
                ("CI95 Low (Mean)", mc_res.ci95_low),
                ("CI95 High (Mean)", mc_res.ci95_high),
            ]

            row_num = 5
            for stat_name, stat_val in mc_stats:
                ws_mc.cell(row=row_num, column=1, value=stat_name)
                val_cell = ws_mc.cell(
                    row=row_num, column=2, value=round(stat_val, 1),
                )
                val_cell.alignment = center
                val_cell.number_format = "0.0"
                # Highlight P99 and Peak above thresholds
                if stat_name in ("P99", "P99.9", "Peak (Max)"):
                    if stat_val >= yellow_threshold:
                        val_cell.fill = yellow_fill
                row_num += 1

            _apply_table_border(ws_mc, 4, row_num - 1, 1, 2)
            _set_column_widths(ws_mc, {"A": 22, "B": 18})

            # Additional info
            row_num += 1
            ws_mc.cell(row=row_num, column=1,
                       value=f"Elapsed: {mc_res.elapsed_sec:.1f} s").font = Font(italic=True)
            ws_mc.cell(row=row_num + 1, column=1,
                       value=f"Seed: {mc_res.seed}").font = Font(italic=True)

        except Exception as exc:  # noqa: BLE001
            logger.warning("Monte Carlo sheet skipped: %s", exc)

    # ── Column widths (main sheet) ────────────────────────────────────────────
    _set_column_widths(ws, {"A": 55, "B": 18, "C": 20})
    for c in range(4, 2 + len(calibrations) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ── Save with locked-file fallback ────────────────────────────────────────
    abs_path = os.path.abspath(output_path)
    try:
        wb.save(abs_path)
    except PermissionError:
        stem, ext = os.path.splitext(abs_path)
        ts = time.strftime("%Y%m%d_%H%M%S")
        fallback = f"{stem}_{ts}{ext}"
        logger.warning(
            "File locked: %s – saving to %s instead.", abs_path, fallback,
        )
        try:
            wb.save(fallback)
            abs_path = fallback
        except OSError as save_err:
            raise ExcelReportError(
                f"Cannot save report: {save_err}"
            ) from save_err
    except OSError as save_err:
        raise ExcelReportError(
            f"Cannot save report: {save_err}"
        ) from save_err

    logger.info("Excel report saved: %s", abs_path)
    return abs_path
