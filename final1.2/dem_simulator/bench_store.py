"""Centralized bench-data store for collaborative simulator training.

Architecture
------------
The store uses an **SQLite database** (``bench_store.db``) on a network
drive or local path.  SQLite provides:

  * **WAL mode** — multiple readers can query concurrently.
  * **Selective I/O** — only the requested project is read, not the entire
    dataset (unlike the legacy JSON approach).
  * **Atomic transactions** — ``BEGIN … COMMIT`` guarantees all-or-nothing.
  * **Zero dependencies** — ``sqlite3`` ships with Python.

Each time a worker completes a bench test they call :func:`upload_bench_result`
to append their measured data.  The simulator queries this store via
:func:`get_fitted_costs` / :func:`get_all_configs` to:

  1. Retrieve pre-fitted ``MicroCosts`` for a known project (RMSE < 5 µs).
  2. Enrich the transfer-learning candidate pool with **real** configs and
     costs from all previously tested projects — greatly improving accuracy
     for *untested* projects.

Default location
----------------
``bench_store.db`` in the centralized bench results folder, or next to
this module as fallback.
Override via :func:`set_store_path` or the ``DEM_BENCH_STORE`` env-var.
"""

from __future__ import annotations

import getpass
import logging
import os
from typing import Any, Dict, List, Optional, Tuple

from dem_simulator.config import ProjectConfig
from dem_simulator.costs import MicroCosts
from dem_simulator.exceptions import SimulatorError
import dem_simulator.bench_store_db as _db

logger = logging.getLogger("dem_simulator")

# ==============================================================================
# Default store path
# ==============================================================================

# Centralized bench results root — same folder as the bench Excel/project data
_CENTRALIZED_DIR = r"\\vt1.vitesco.com\SMT\did01146\Aggr_info\ERRM_Error_Management\00_Aggregate_Generic\FinalTests_Results\WCS\bench_results"
_CENTRALIZED_STORE_DB = os.path.join(_CENTRALIZED_DIR, "bench_store.db")

# Priority: env-var → centralized folder → fallback next to this module
_FALLBACK_STORE_DB = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "bench_store.db"
)


def _resolve_default_store() -> str:
    """Determine the default DB path, creating the centralized dir if needed.

    Always prefers ``d:\\casdev\\td5\\BM\\bench_results\\bench_store.db``.
    If the directory cannot be created (e.g. network drive offline),
    falls back to the module directory.
    """
    try:
        os.makedirs(_CENTRALIZED_DIR, exist_ok=True)
        return _CENTRALIZED_STORE_DB
    except OSError:
        return _FALLBACK_STORE_DB


_store_path: str = os.environ.get("DEM_BENCH_STORE", _resolve_default_store())
_seeded_paths: set[str] = set()


def get_store_path() -> str:
    """Return the current bench-store file path (SQLite .db)."""
    return _store_path


def seed_bench_store(store_path: Optional[str] = None, *, force: bool = False) -> int:
    """Populate the SQLite bench store with validated reference projects.

    The operation is idempotent by default: existing projects are not
    overwritten unless ``force=True``.
    """
    from dem_simulator.seeds import SEED_ENTRIES
    from dem_simulator.config import FrfBlockConfig

    path = store_path or _store_path
    written = 0

    for entry in SEED_ENTRIES:
        key = entry["project_name"].upper()
        if not force and _db.get_fitted_costs(path, key) is not None:
            continue

        cfg_kwargs = dict(entry["cfg_kwargs"])
        frf_raw = cfg_kwargs.pop("FrfBlocks", [])
        cfg = ProjectConfig(
            FrfBlocks=[FrfBlockConfig(**block) for block in frf_raw],
            **cfg_kwargs,
        )
        costs = MicroCosts(**entry["cost_kwargs"])
        _db.upload_bench_result(
            db_path=path,
            project_name=key,
            bench=entry["bench"],
            cfg=cfg,
            fitted_costs=costs,
            fit_rmse=entry.get("fit_rmse", 0.0),
            calibrations=entry.get("calibrations"),
            uploaded_by=f"seed:{entry.get('source', 'seeds.py')}",
        )
        written += 1

    return written


def _ensure_seeded(store_path: Optional[str] = None) -> None:
    """Seed the *default* DB once per process, without failing the caller.

    Only the path currently configured as :data:`_store_path` (i.e. the
    centralized production store) is auto-seeded.  Tests and other
    callers that pass an explicit ``store_path`` get a clean database —
    they are expected to populate it themselves.
    """
    target = os.path.abspath(store_path or _store_path)
    default = os.path.abspath(_store_path)
    if target != default:
        # Explicit store path → no automatic seeding (keeps unit tests
        # deterministic and prevents accidental cross-contamination).
        return
    if target in _seeded_paths:
        return
    try:
        seed_bench_store(target)
    except Exception as exc:
        logger.warning("[Seed] Auto-seed failed for %s: %s", target, exc)
    finally:
        _seeded_paths.add(target)


def set_store_path(path: str) -> None:
    """Set the active bench-store database path.
    
    This updates the module-level _store_path variable used for all subsequent
    store operations. Useful for UI/CLI tools that need to switch store locations
    at runtime.
    
    Parameters
    ----------
    path : str
        Absolute path to the SQLite database file (*.db).
        
    Raises
    ------
    ValueError
        If path does not end with .db
    """
    global _store_path
    if not path.lower().endswith(".db"):
        raise ValueError(f"Store path must end with .db, got: {path}")
    _store_path = os.path.abspath(path)
    _seeded_paths.discard(_store_path)
    logger.info(f"[BenchStore] Active store path changed to: {_store_path}")


def clear_store(store_path: Optional[str] = None, *, force: bool = False) -> None:
    """DESTRUCTIVE: Delete all projects from the bench store.
    
    WARNING: This cannot be undone. All bench data, configs, and fitted costs
    in the store will be permanently deleted.
    
    Parameters
    ----------
    store_path : str, optional
        Override the store file location. Defaults to the active store.
    force : bool
        If True, skips safety confirmation prompt (use with caution).
    
    Raises
    ------
    FileNotFoundError
        If the store file does not exist.
    """
    import warnings
    path = store_path or _store_path
    
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Bench-store DB not found at {path}")
    
    if not force:
        warnings.warn(
            f"About to DELETE ALL data from: {path}\n"
            "This is destructive and cannot be undone.\n"
            "Call again with force=True to proceed.",
            UserWarning
        )
        return
    
    try:
        _db._clear_all_data(path)
        logger.warning("[BenchStore] All data cleared from %s", path)
    except Exception as exc:
        logger.error("[BenchStore] Failed to clear store: %s", exc)
        raise


# ==============================================================================
# Serialisation helpers  (kept for parse_bench_excel / ingest path only)
# ==============================================================================

def _normalise_key(project_name: str) -> str:
    """Normalise project name to upper-case, preserving the full name.

    Examples::

        "PROJ3_0U0_P16_624" → "PROJ3_0U0_P16_624"
        "  proj1  "          → "PROJ1"
        "FOH12_0U0"          → "FOH12_0U0"
    """
    return project_name.strip().upper()


# ==============================================================================
# Public API — Upload
# ==============================================================================

def upload_bench_result(
    project_name: str,
    bench: Dict[str, List[int]],
    cfg: Optional[ProjectConfig] = None,
    *,
    fitted_costs: Optional[MicroCosts] = None,
    fit_rmse: float = 0.0,
    calibrations: Optional[List[Tuple[int, int]]] = None,
    uploaded_by: str = "",
    store_path: Optional[str] = None,
) -> str:
    """Upload bench results for a project to the SQLite store.

    If *fitted_costs* is not provided but *cfg* is, a :func:`transfer_fit`
    is run automatically and the resulting costs are stored.

    Parameters
    ----------
    project_name : str
        Project key (e.g. ``"PROJ3"`` or ``"PROJ3_0U0_P16_624"``).
    bench : dict
        ``{ "scenario_1": [v1..v6], "scenario_2": [...], "scenario_3": [...] }``
    cfg : ProjectConfig, optional
        Extracted project configuration (stored for transfer-learning).
    fitted_costs : MicroCosts, optional
        Pre-fitted costs.  If None and *cfg* is given, auto-fit runs.
    fit_rmse : float
        RMSE of the fitted costs (stored for diagnostics).
    calibrations : list, optional
        Calibration pairs used (default: standard 6).
    uploaded_by : str
        Username / workstation identifier.
    store_path : str, optional
        Override the store file location for this call.

    Returns
    -------
    str
        The normalised project key under which the data was stored.
    """
    path = store_path or _store_path
    _ensure_seeded(path)
    # _ensure_migrated()
    return _db.upload_bench_result(
        db_path=path,
        project_name=project_name,
        bench=bench,
        cfg=cfg,
        fitted_costs=fitted_costs,
        fit_rmse=fit_rmse,
        calibrations=calibrations,
        uploaded_by=uploaded_by,
    )


# ==============================================================================
# Public API — Query
# ==============================================================================

def get_fitted_costs(
    project_name: str,
    store_path: Optional[str] = None,
) -> Optional[MicroCosts]:
    """Retrieve pre-fitted MicroCosts for a project, or None.

    Exact match, then underscore-anchored prefix match.  If multiple
    prefix matches, the first alphabetically is used and a warning is
    logged.
    """
    path = store_path or _store_path
    _ensure_seeded(path)
    # _ensure_migrated()
    return _db.get_fitted_costs(path, project_name)


def get_bench_data(
    project_name: str,
    store_path: Optional[str] = None,
) -> Optional[Dict[str, List[int]]]:
    """Retrieve raw bench data for a project, or None."""
    path = store_path or _store_path
    _ensure_seeded(path)
    # _ensure_migrated()
    return _db.get_bench_data(path, project_name)


def get_all_configs(
    store_path: Optional[str] = None,
) -> Dict[str, ProjectConfig]:
    """Return all stored ProjectConfigs (for enriching transfer-learning).

    Returns
    -------
    dict
        ``{ "PROJ3": ProjectConfig, "PROJ6": ProjectConfig, ... }``
        Only projects that have a stored ``config_summary`` are included.
    """
    path = store_path or _store_path
    _ensure_seeded(path)
    # _ensure_migrated()
    return _db.get_all_configs(path)


def get_all_fitted_costs(
    store_path: Optional[str] = None,
) -> Dict[str, MicroCosts]:
    """Return all stored fitted costs (for enriching transfer-learning).

    Returns
    -------
    dict
        ``{ "PROJ3": MicroCosts, ... }``
    """
    path = store_path or _store_path
    _ensure_seeded(path)
    # _ensure_migrated()
    return _db.get_all_fitted_costs(path)


def list_projects(
    store_path: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Return a summary list of all projects in the store.

    Each entry has: ``key``, ``has_bench``, ``has_costs``, ``fit_rmse``,
    ``uploaded_by``, ``uploaded_at``.
    """
    path = store_path or _store_path
    # _ensure_migrated()
    return _db.list_projects(path)


# ==============================================================================
# Internal helpers
# ==============================================================================

def _key_variants(project_name: str) -> List[str]:
    """Generate candidate keys for lookup.

    Returns the full upper-cased name plus a short prefix fallback
    (for backwards compatibility with legacy data stored under short keys).
    """
    key = project_name.strip().upper()
    variants = [key]
    prefix = key.split("_")[0]
    if prefix != key and prefix not in variants:
        variants.append(prefix)
    return variants


# ==============================================================================
# Excel parser — read bench values from a WCS results spreadsheet
# ==============================================================================

def parse_bench_excel(
    excel_path: str,
    sheet_name: str = "Runtime_DemMainFct",
) -> Dict[str, Any]:
    """Extract bench-measured runtime values from a RuntimeMeasureReduction Excel.

    Reads the standard ``RuntimeMeasureReduction.xlsx`` template used by
    the bench measurement process.  The target sheet (default
    ``Runtime_DemMainFct``) has this layout:

    ::

        Row  1: "Add here project name…"  |  "Value in project:"  |  …
        Row  2: "NrFmy in project:"        |  <NrFmy int>          |  …
        Row  3: "Icsp_Dem_Fmy_NrClcFmyEveAsyn:"  |  …
        Row  4: "Icsp_Dem_Fmy_NrClcFmyPost:"     |  …
        Row  5-8: (MBT, Core, Task, …)
        Row  9: (blank C1)  |  "Calibration 1:"  |  …  |  "Calibration 6:"
        Row 10: "… NrClcFmyEveAsyn/…Post… Worst case scenarios:"  |  "20/10" | …
        Row 11: Scenario 1 values (C2..C7)  — integers (µs)
        Row 12: Scenario 2 values (C2..C7)
        Row 13: Scenario 3 values (C2..C7)

    The parser also extracts the calibration tuples (asyn, post) from Row 10
    and the NrFmy count from Row 2.

    Parameters
    ----------
    excel_path : str
        Path to the ``.xlsx`` file (``RuntimeMeasureReduction.xlsx``).
    sheet_name : str
        Name of the worksheet to read.  Defaults to ``Runtime_DemMainFct``.
        Falls back to the first sheet if not found.

    Returns
    -------
    dict
        ``{ "bench":       { "scenario_1": [v1..v6], … },``
        ``  "calibrations": [(asyn,post), …],``
        ``  "nr_fmy":      int }``

    Raises
    ------
    SimulatorError
        If the file cannot be read or the expected table is not found.
    """
    try:
        import openpyxl
    except ImportError:
        raise SimulatorError(
            "openpyxl is required to parse Excel files.  "
            "Install with:  pip install openpyxl"
        )

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as exc:
        raise SimulatorError(f"Cannot open Excel file {excel_path!r}: {exc}")

    # ---- locate the sheet ----
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Try common alternatives
        for alt in ("Runtime_DemMainFct", "WCS Results", "Runtime"):
            if alt in wb.sheetnames:
                ws = wb[alt]
                break
        else:
            ws = wb.worksheets[0]
        logger.debug(
            "[BenchStore] Sheet '%s' not found — using '%s'.",
            sheet_name, ws.title,
        )

    # ---- find the calibration header row (contains "Calibration 1:") ----
    cal_header_row = None
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
        for cell in row:
            val = str(cell.value or "").strip().lower()
            if "calibration 1" in val:
                cal_header_row = cell.row
                break
        if cal_header_row is not None:
            break

    if cal_header_row is None:
        wb.close()
        raise SimulatorError(
            f"Could not find calibration header row in '{ws.title}' of "
            f"{excel_path!r}.  Expected a row containing 'Calibration 1:'"
        )

    # ---- identify calibration columns (those with "Calibration N:") ----
    cal_cols: List[int] = []  # 1-based column numbers
    for cell in ws[cal_header_row]:
        val = str(cell.value or "").strip().lower()
        if "calibration" in val:
            cal_cols.append(cell.column)

    # Fallback: columns B..G (2..7)
    if len(cal_cols) < 2:
        cal_cols = list(range(2, 8))

    n_cals = len(cal_cols)

    # ---- parse calibration tuples (asyn/post) from next row ----
    asyn_post_row = cal_header_row + 1
    calibrations: List[tuple] = []
    for col in cal_cols:
        raw = str(ws.cell(row=asyn_post_row, column=col).value or "")
        calibrations.append(_parse_asyn_post(raw))

    # ---- find scenario data rows (3 rows after asyn/post row) ----
    scenario_keys = ["scenario_1", "scenario_2", "scenario_3"]
    bench: Dict[str, List[int]] = {}
    for si, skey in enumerate(scenario_keys):
        data_row = asyn_post_row + 1 + si
        row_vals: List[int] = []
        for col in cal_cols:
            raw = ws.cell(row=data_row, column=col).value
            row_vals.append(_parse_cell_int(raw))
        bench[skey] = row_vals

    # ---- extract NrFmy from row 2, column B ----
    nr_fmy = 0
    core_name = ""
    task_name = ""
    for r in range(1, cal_header_row):
        c1 = str(ws.cell(row=r, column=1).value or "").lower()
        if "nrfmy" in c1.replace(" ", "").replace("_", ""):
            nr_fmy = _parse_cell_int(ws.cell(row=r, column=2).value)
        if "core in project" in c1 or c1.strip() == "core:":
            core_name = str(ws.cell(row=r, column=2).value or "").strip()
        if "task in project" in c1 or c1.strip() == "task:":
            task_name = str(ws.cell(row=r, column=2).value or "").strip()

    task_lower = task_name.lower()
    is_shared_core = ("systrig" in task_lower) or ("shared" in task_lower)

    wb.close()

    # ---- validate ----
    total = sum(v for row in bench.values() for v in row)
    if total == 0:
        raise SimulatorError(
            f"All values read from {excel_path!r} are zero.  "
            "Check that the correct sheet and cell range contain bench data."
        )

    logger.info(
        "[BenchStore] Parsed bench Excel %s (NrFmy=%d, %d calibrations):",
        os.path.basename(excel_path), nr_fmy, n_cals,
    )
    for skey in scenario_keys:
        logger.info("  %s: %s", skey, bench[skey])

    return {
        "bench": bench,
        "calibrations": calibrations,
        "nr_fmy": nr_fmy,
        "core": core_name,
        "task": task_name,
        "is_shared_core": is_shared_core,
    }


def _parse_cell_int(raw) -> int:
    """Convert an Excel cell value to int.

    Handles several real-world bench cell formats:

    * Pure numbers (``int`` / ``float``) → rounded to ``int``.
    * Single numeric string, possibly with units / non-breaking spaces
      (``"498"``, ``"596us"``, ``"498\\xa0"``) → numeric value.
    * Multi-measurement strings using ``/`` or ``-`` separators
      (``"596/594"``, ``"596 / 594"``, ``"595-600"``) → **average** of
      the two values, rounded.  This is essential for projects that
      report ``min/max`` or ``run1/run2`` in the same cell — otherwise
      the slash would be stripped and ``"596/594"`` would be parsed as
      ``596594`` (catastrophic 1000x error).
    * Empty / blank / non-parseable → ``0``.
    """
    import re

    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(round(raw))

    text = str(raw).strip()
    if not text:
        return 0

    # Extract all numeric tokens (handles units like 'us', NBSP, newlines)
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    if not nums:
        return 0

    try:
        values = [float(n) for n in nums]
    except (ValueError, TypeError):
        return 0

    # Heuristic: multiple numeric tokens with a separator → multi-measurement
    # cell (e.g. "596/594", "595-600", "min:596 max:594").  Average them.
    if len(values) >= 2 and re.search(r"[\/\-,;]", text):
        return int(round(sum(values) / len(values)))

    # Single numeric token (or repeated identical ones) → first value.
    return int(round(values[0]))


def _parse_asyn_post(raw: str) -> tuple:
    """Parse a calibration cell like ``'20/\\n10\\n'`` → ``(20, 10)``.

    The bench Excel stores NrClcFmyEveAsyn / NrClcFmyPost as multi-line
    strings separated by ``/``.  This extracts the two integers.
    """
    import re
    nums = re.findall(r"\d+", str(raw))
    if len(nums) >= 2:
        return (int(nums[0]), int(nums[1]))
    elif len(nums) == 1:
        return (int(nums[0]), int(nums[0]))
    return (0, 0)


# ==============================================================================
# Folder-based ingestion — scan a shared folder for Excel + project pairs
# ==============================================================================

# Default bench results root (can be overridden)
_BENCH_RESULTS_ROOT = r"\\vt1.vitesco.com\SMT\did01146\Aggr_info\ERRM_Error_Management\00_Aggregate_Generic\FinalTests_Results\WCS\bench_results"

# Known bench Excel filenames (case-insensitive matching)
_BENCH_EXCEL_NAMES = (
    "runtimemeasurereduction.xlsx",
    "results_wcs.xlsx",
    "runtime_demmainfunction.xlsx",
)


def ingest_single_project(
    excel_path: str,
    project_path: str,
    project_name: str = "",
    *,
    store_path: Optional[str] = None,
    cpu_clock_mhz: float = 300.0,
) -> str:
    """Ingest a single project from an Excel file + project directory.

    This is the main function called from the UI when a worker uploads
    a bench result.

    Parameters
    ----------
    excel_path : str
        Path to the WCS results Excel file with bench measurements.
    project_path : str
        Path to the built project directory (containing ``icsp_dem_cnf.c``
        somewhere in its tree).
    project_name : str
        Project name override.  If empty, inferred from *project_path*.
    store_path : str, optional
        Override bench store location.
    cpu_clock_mhz : float
        CPU clock for config extraction.

    Returns
    -------
    str
        The normalised project key under which data was stored.
    """
    # Parse bench values from Excel
    parsed = parse_bench_excel(excel_path)
    bench = parsed["bench"]

    logger.info(
        "[BenchStore] Calibrations from Excel: %s  (NrFmy=%d)",
        parsed["calibrations"], parsed["nr_fmy"],
    )

    # Find icsp_dem_cnf.c in the project tree
    cnf_c = _find_cnf_c(project_path)
    if cnf_c is None:
        raise SimulatorError(
            f"Could not find icsp_dem_cnf.c under {project_path!r}.\n"
            "Make sure you point to the built project directory."
        )

    # Infer project name from directory if not given
    if not project_name:
        project_name = os.path.basename(os.path.normpath(project_path))

    # Find PTU directory
    ptu_dir = _find_ptu_dir(project_path)

    # Extract config
    from dem_simulator.extractor import load_project_from_c
    project_def = load_project_from_c(
        cnf_c_path=cnf_c,
        project_name=project_name,
        ptu_dir=ptu_dir,
        cpu_clock_mhz=cpu_clock_mhz,
    )

    # Upload (auto-fits costs)
    key = upload_bench_result(
        project_name=project_name,
        bench=bench,
        cfg=project_def.default_config,
        calibrations=parsed.get("calibrations"),
        uploaded_by=(
            os.environ.get("USERNAME")
            or os.environ.get("USER")
            or getpass.getuser()
        ),
        store_path=store_path,
    )

    logger.info(
        "[BenchStore] Ingested project '%s' from Excel=%s, Project=%s",
        key, os.path.basename(excel_path), project_path,
    )
    return key


def _find_cnf_c(root: str) -> Optional[str]:
    """Recursively find ``icsp_dem_cnf.c`` under *root*.

    When multiple copies exist (e.g. under ``bld/`` and ``proc/``), the
    copy under ``proc/`` is preferred because ``bld/`` is a build-cache
    duplicate that lacks sibling artefacts like ``icsp_dem_genr.xml``.

    The ``bld`` subtree is skipped entirely during the initial walk to
    avoid scanning thousands of cached build files.  If nothing is found
    outside ``bld``, a second pass searches inside it as a fallback.
    """
    target = "icsp_dem_cnf.c"
    _SKIP_DIRS = {"bld", ".git", "__pycache__"}

    # First pass — skip bld directories
    best: Optional[str] = None
    for dirpath, dirnames, filenames in os.walk(root):
        # Prune unwanted subtrees in-place
        dirnames[:] = [d for d in dirnames if d.lower() not in _SKIP_DIRS]
        for fn in filenames:
            if fn.lower() == target:
                candidate = os.path.join(dirpath, fn)
                # Prefer proc\...\out\ path
                if "proc" in candidate.lower() and \
                   candidate.lower().endswith(os.sep + "out" + os.sep + target):
                    return candidate
                if best is None:
                    best = candidate
    if best is not None:
        return best

    # Fallback — search inside bld if nothing found above
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower() == target:
                return os.path.join(dirpath, fn)
    return None


def _find_ptu_dir(root: str) -> Optional[str]:
    """Recursively find a directory containing ``icsp_dem_cnf_ptu*.h``."""
    for dirpath, _, filenames in os.walk(root):
        ptu_files = [f for f in filenames
                     if f.lower().startswith("icsp_dem_cnf_ptu")
                     and f.lower().endswith(".h")]
        if ptu_files:
            return dirpath
    return None


def _ingest_single(
    xlsx_path: str,
    *,
    store_path: Optional[str] = None,
    cpu_clock_mhz: float = 300.0,
) -> str:
    """Ingest a single Excel + project pair found during folder scan."""
    project_dir = os.path.dirname(xlsx_path)
    project_name = os.path.basename(project_dir)
    return ingest_single_project(
        excel_path=xlsx_path,
        project_path=project_dir,
        project_name=project_name,
        store_path=store_path,
        cpu_clock_mhz=cpu_clock_mhz,
    )
